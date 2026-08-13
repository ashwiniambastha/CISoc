"""
Shared append-to-log used by both the batch runner and the dashboard.

Every row goes to two places:
  1. logs/run_log.jsonl  -- append-only, written first, verbatim, crash-proof
  2. logs/run_log.xlsx   -- the analysis sheet

The JSONL is the authoritative record (Appendix C.7.7): Excel can be locked by
the desktop app or corrupted by a kill mid-write, a line of JSON cannot.
"""

from __future__ import annotations

import json
import re
import threading
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Set, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

import config

# Excel refuses control characters and caps a cell at 32,767 characters.
_ILLEGAL = re.compile(r"[\000-\010\013\014\016-\037]")
_CELL_LIMIT = 32_000

_WIDTHS = {
    "response_text": 80,
    "prompt_text": 46,
    "record_id": 12,
    "timestamp": 22,
    "prompt_id": 12,
    "model_name": 12,
    "model_id_sent": 26,
    "model_id_returned": 26,
    "bank_sha256": 18,
}


# --------------------------------------------------------------------------


def _clean(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    text = str(value)
    text = _ILLEGAL.sub("", text)
    if len(text) > _CELL_LIMIT:
        text = text[:_CELL_LIMIT] + f"...[TRUNCATED FOR EXCEL, full text in run_log.jsonl]"
    return text


def ensure_workbook(path: Path = None) -> Path:
    path = Path(path or config.EXCEL_LOG_PATH)
    if path.exists():
        # A schema change would otherwise shift every value one column left
        # from the row it was appended on, silently.
        header = [c.value for c in load_workbook(path, read_only=True)["elicitations"][1]]
        if header != config.COLUMNS:
            missing = [c for c in config.COLUMNS if c not in header]
            extra = [c for c in header if c not in config.COLUMNS]
            raise RuntimeError(
                f"\n{path.name} was written under a different schema.\n"
                f"  new columns: {missing or 'none'}\n"
                f"  dropped:     {extra or 'none'}\n"
                f"Rename the old log (e.g. run_log_pre_frame.xlsx, and the .jsonl\n"
                f"alongside it) and let a fresh one be created. Rows from before a\n"
                f"schema or frame change are a separate run and should not be pooled\n"
                f"with what follows."
            )
        return path
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "elicitations"
    ws.append(config.COLUMNS)
    for i, col in enumerate(config.COLUMNS, start=1):
        cell = ws.cell(row=1, column=i)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = _WIDTHS.get(col, 15)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(config.COLUMNS))}1"
    wb.save(path)
    return path


# Both models are called concurrently from the dashboard; two threads opening
# the same workbook would lose rows. One writer at a time.
_WRITE_LOCK = threading.Lock()


def append_rows(
    rows: Iterable[Dict[str, Any]],
    excel_path: Path = None,
    jsonl_path: Path = None,
) -> int:
    """Append rows to the JSONL first, then the Excel sheet. Thread-safe."""
    rows = [r for r in rows if r]
    if not rows:
        return 0

    excel_path = Path(excel_path or config.EXCEL_LOG_PATH)
    jsonl_path = Path(jsonl_path or config.JSONL_LOG_PATH)
    jsonl_path.parent.mkdir(parents=True, exist_ok=True)

    with _WRITE_LOCK:
        return _append_locked(rows, excel_path, jsonl_path)


def _append_locked(rows, excel_path: Path, jsonl_path: Path) -> int:
    with jsonl_path.open("a", encoding="utf-8") as fh:
        for row in rows:
            fh.write(json.dumps(row, ensure_ascii=False, default=str) + "\n")
            fh.flush()

    ensure_workbook(excel_path)
    try:
        wb = load_workbook(excel_path)
        ws = wb["elicitations"]
        for row in rows:
            ws.append([_clean(row.get(col, "")) for col in config.COLUMNS])
        wb.save(excel_path)
    except PermissionError:
        # Almost always: the file is open in Excel. Never lose the data.
        sidecar = excel_path.with_name(
            f"{excel_path.stem}_recovered_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        )
        ensure_workbook(sidecar)
        wb = load_workbook(sidecar)
        ws = wb["elicitations"]
        for row in rows:
            ws.append([_clean(row.get(col, "")) for col in config.COLUMNS])
        wb.save(sidecar)
        print(
            f"  [warn] {excel_path.name} is locked (open in Excel?). "
            f"Wrote {len(rows)} row(s) to {sidecar.name} instead. "
            "The JSONL is complete either way."
        )
    return len(rows)


append_row = lambda row, **kw: append_rows([row], **kw)  # noqa: E731


# --------------------------------------------------------------------------
# Reading back
# --------------------------------------------------------------------------


def read_all(jsonl_path: Path = None) -> List[Dict[str, Any]]:
    """Read the JSONL log (the authoritative record)."""
    jsonl_path = Path(jsonl_path or config.JSONL_LOG_PATH)
    if not jsonl_path.exists():
        return []
    out = []
    with jsonl_path.open("r", encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if line:
                try:
                    out.append(json.loads(line))
                except json.JSONDecodeError:
                    continue
    return out


def read_tail(n: int = 10, jsonl_path: Path = None) -> List[Dict[str, Any]]:
    return read_all(jsonl_path)[-n:][::-1]


def completed_keys(jsonl_path: Path = None) -> Set[Tuple]:
    """Successful (prompt_id, model_name, condition, replicate_index) keys.

    Used by the batch runner to resume without re-eliciting. Only rows with
    source='batch' count: a dashboard run is a different event class.
    """
    keys = set()
    for row in read_all(jsonl_path):
        if row.get("status") == "success" and row.get("source") == "batch":
            keys.add(
                (
                    row.get("prompt_id"),
                    row.get("model_name"),
                    row.get("condition"),
                    int(row.get("replicate_index") or 0),
                )
            )
    return keys


def summarise(jsonl_path: Path = None) -> Dict[str, Any]:
    rows = read_all(jsonl_path)
    ok = [r for r in rows if r.get("status") == "success"]
    return {
        "attempts": len(rows),
        "success": len(ok),
        "failed": len(rows) - len(ok),
        "truncated": sum(1 for r in ok if r.get("truncated") in (True, "True")),
        "by_model": {
            m: sum(1 for r in ok if r.get("model_name") == m) for m in config.MODELS
        },
        "total_tokens": sum(
            int(r["total_tokens"]) for r in ok if str(r.get("total_tokens", "")).isdigit()
        ),
    }


if __name__ == "__main__":
    import config as _c

    test_row = {c: "" for c in _c.COLUMNS}
    test_row.update(
        {
            "record_id": "test-0001",
            "timestamp": datetime.now().isoformat(timespec="seconds"),
            "source": "self_test",
            "prompt_id": "K-K1a-HI",
            "language": "HI",
            "model_name": "nemotron",
            "condition": "default",
            "replicate_index": 1,
            "temperature": "unset",
            "status": "success",
            "response_text": "क्या परिवार एक व्यक्ति है? — self test row, safe to delete.",
            "total_tokens": 42,
        }
    )
    append_rows([test_row])
    print("wrote 1 test row")
    print(json.dumps(read_tail(1)[0], ensure_ascii=False, indent=2)[:400])
    print(summarise())
