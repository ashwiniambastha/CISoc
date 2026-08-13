"""
Prove the log is intact, and repair display formatting on older rows.

    python verify_log.py             # check
    python verify_log.py --rebuild   # regenerate run_log.xlsx from run_log.jsonl

Checks every response_text cell in run_log.xlsx against the verbatim string in
run_log.jsonl and reports any that differ. Also re-applies wrap-text so long
responses display in full rather than as a single clipped line.

--rebuild regenerates the spreadsheet from the JSONL, which is what to run
after the column order or schema changes. The previous file is kept as
run_log_prev.xlsx. Nothing is re-elicited.
"""

import sys

from openpyxl import load_workbook
from openpyxl.styles import Alignment

import config
import excel_logger

WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")


def main() -> None:
    if "--rebuild" in sys.argv:
        n = excel_logger.rebuild_excel()
        print(f"rebuilt {config.EXCEL_LOG_PATH.name} from the JSONL: {n} rows, "
              f"{len(config.EXCEL_COLUMNS)} columns")
        print("previous file kept as run_log_prev.xlsx\n")
    rows = excel_logger.read_all()
    if not rows:
        print("log is empty")
        return

    wb = load_workbook(config.EXCEL_LOG_PATH)
    ws = wb["elicitations"]
    header = [c.value for c in ws[1]]
    ri = header.index("response_text") + 1
    ci = header.index("record_id") + 1

    by_id = {
        ws.cell(row=i, column=ci).value: (i, ws.cell(row=i, column=ri).value or "")
        for i in range(2, ws.max_row + 1)
    }

    mismatches, missing, longest = [], [], 0
    for r in rows:
        longest = max(longest, len(r["response_text"]))
        found = by_id.get(r["record_id"])
        if found is None:
            missing.append(r["record_id"])
        elif found[1] != r["response_text"]:
            mismatches.append((r["record_id"], len(r["response_text"]), len(found[1])))

    print(f"rows in jsonl        : {len(rows)}")
    print(f"rows in xlsx         : {ws.max_row - 1}")
    print(f"longest response     : {longest} chars  (Excel's limit is 32,767)")
    print(f"missing from xlsx    : {len(missing)}")
    print(f"text mismatches      : {len(mismatches)}")
    for rid, a, b in mismatches[:10]:
        print(f"   {rid}: jsonl {a} chars vs xlsx {b} chars")

    # Re-apply wrap formatting to every data row.
    for i in range(2, ws.max_row + 1):
        for j, col in enumerate(config.EXCEL_COLUMNS, start=1):
            ws.cell(row=i, column=j).alignment = (
                WRAP if col in excel_logger._WRAP_COLUMNS else TOP
            )
        ws.row_dimensions[i].height = None  # let Excel auto-fit
    wb.save(config.EXCEL_LOG_PATH)
    print(f"\nwrap-text re-applied to {ws.max_row - 1} rows")

    # Replicate coverage, per cell.
    cells = {}
    for r in rows:
        key = (r.get("prompt_id"), r.get("model_name"), r.get("condition"),
               str(r.get("temperature")))
        cells.setdefault(key, []).append(r.get("replicate_index"))
    dupes = {k: v for k, v in cells.items() if len(v) != len(set(v))}
    print(f"\ncells with repeated replicate_index: {len(dupes)}")
    for k, v in list(dupes.items())[:8]:
        print(f"   {k} -> {sorted(v)}")

    if not mismatches and not missing:
        print("\nOK: every response in the spreadsheet is byte-identical to the "
              "verbatim JSONL record.")


if __name__ == "__main__":
    main()
